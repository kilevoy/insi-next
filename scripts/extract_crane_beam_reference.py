from __future__ import annotations

import json
import os
import re
import sys
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

warnings.filterwarnings(
    'ignore',
    message='Data Validation extension is not supported and will be removed',
    category=UserWarning,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
CRANE_BEAM_REFERENCE_PATH = REPO_ROOT / 'src/domain/crane-beam/model/crane-beam-reference.generated.ts'

WORKBOOK_ENV_VAR = 'CRANE_BEAM_REFERENCE_WORKBOOK'
WORKBOOK_CANDIDATES = [
    Path(os.environ[WORKBOOK_ENV_VAR]) if os.environ.get(WORKBOOK_ENV_VAR) else None,
    REPO_ROOT / 'подбор прокатной подкрановой балки v2.0.xlsx',
    REPO_ROOT.parent / 'подбор прокатной подкрановой балки v2.0.xlsx',
    Path('C:/Users/RAN/Downloads/подбор прокатной подкрановой балки v2.0.xlsx'),
]

TS_EXPORT_PATTERN = re.compile(r'export const (\w+) = (.*? as const;)', re.S)


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def coerce_float(value: Any) -> float:
    if is_number(value):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(',', '.')
        if normalized:
            return float(normalized)
    raise ValueError(f'Expected numeric value, got {value!r}')


def find_workbook_path() -> Path | None:
    for candidate in WORKBOOK_CANDIDATES:
        if candidate and candidate.exists():
            return candidate
    return None


def load_snapshot_exports(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}

    text = path.read_text(encoding='utf-8')
    exports: dict[str, Any] = {}

    for export_name, block in TS_EXPORT_PATTERN.findall(text):
        json_text = block[: -len(' as const;')]
        exports[export_name] = json.loads(json_text)

    return exports


def extract_candidate_catalog(ws_calculation, expected_count: int | None) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []

    for row in range(3, ws_calculation.max_row + 1):
        profile = ws_calculation[f'H{row}'].value
        area_cm2 = ws_calculation[f'P{row}'].value
        unit_mass = ws_calculation[f'Q{row}'].value
        resistance = ws_calculation[f'G{row}'].value

        if not isinstance(profile, str) or not profile.strip():
            continue
        if not is_number(area_cm2) or not is_number(unit_mass) or not is_number(resistance):
            continue

        excluded = str(ws_calculation[f'E{row}'].value or '').strip() == '-'

        candidates.append(
            {
                'ordinal': len(candidates) + 1,
                'row': row,
                'profile': profile.strip(),
                'excluded': excluded,
                'ryMpa': coerce_float(resistance),
                'hMm': coerce_float(ws_calculation[f'I{row}'].value),
                'bMm': coerce_float(ws_calculation[f'J{row}'].value),
                'webThicknessMm': coerce_float(ws_calculation[f'K{row}'].value),
                'flangeThicknessMm': coerce_float(ws_calculation[f'L{row}'].value),
                'webHeightMm': coerce_float(ws_calculation[f'M{row}'].value),
                'flangeOverhangMm': coerce_float(ws_calculation[f'N{row}'].value),
                'radiusMm': coerce_float(ws_calculation[f'O{row}'].value),
                'areaCm2': coerce_float(area_cm2),
                'unitMassKgPerM': coerce_float(unit_mass),
                'ixCm4': coerce_float(ws_calculation[f'R{row}'].value),
                'wxCm3': coerce_float(ws_calculation[f'S{row}'].value),
                'sxCm3': coerce_float(ws_calculation[f'T{row}'].value),
                'ixRadiusCm': coerce_float(ws_calculation[f'U{row}'].value),
                'iyCm4': coerce_float(ws_calculation[f'V{row}'].value),
                'wyCm3': coerce_float(ws_calculation[f'W{row}'].value),
                'syCm3': coerce_float(ws_calculation[f'X{row}'].value),
                'iyRadiusCm': coerce_float(ws_calculation[f'Y{row}'].value),
                'torsionItCm4': coerce_float(ws_calculation[f'Z{row}'].value),
                'warpingIwCm6': coerce_float(ws_calculation[f'AA{row}'].value),
                'sectionWcm2': coerce_float(ws_calculation[f'AB{row}'].value),
                'ifCm4': coerce_float(ws_calculation[f'AC{row}'].value),
                'i1fCm4': coerce_float(ws_calculation[f'AD{row}'].value),
            }
        )

    if expected_count is not None and expected_count != len(candidates):
        raise ValueError(
            f'Expected {expected_count} crane beam candidates, extracted {len(candidates)} from {ws_calculation.title!r}'
        )

    return candidates


def extract_reference_exports(workbook_path: Path, snapshot_exports: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    ws_calculation = workbook[workbook.sheetnames[1]]

    expected_count = None
    if 'craneBeamCandidateCatalog' in snapshot_exports:
      expected_count = len(snapshot_exports['craneBeamCandidateCatalog'])

    return {
        'craneBeamCandidateCatalog': extract_candidate_catalog(ws_calculation, expected_count),
    }


def render_module(exports_map: dict[str, Any], workbook_path: Path | None) -> str:
    if workbook_path is None:
        header = [
            '// Rebuilt from the checked-in crane beam reference snapshot.',
            '// Workbook-backed extraction was skipped because no source workbook was found.',
        ]
    else:
        header = [
            '// Rebuilt from workbook-backed crane beam references.',
            f'// Source workbook: {workbook_path.name}',
        ]

    lines = [*header, '']
    for export_name, value in exports_map.items():
        lines.append(f'export const {export_name} = {json.dumps(value, ensure_ascii=False, indent=2)} as const;')
        lines.append('')

    return '\n'.join(lines).rstrip() + '\n'


def main() -> int:
    snapshot_exports = load_snapshot_exports(CRANE_BEAM_REFERENCE_PATH)
    workbook_path = find_workbook_path()

    if workbook_path is None:
        if not snapshot_exports:
            raise FileNotFoundError(
                'Unable to locate crane beam workbook and no checked-in reference snapshot exists.'
            )
        rendered = render_module(snapshot_exports, workbook_path=None)
    else:
        exports_map = extract_reference_exports(workbook_path, snapshot_exports)
        rendered = render_module(exports_map, workbook_path=workbook_path)

    CRANE_BEAM_REFERENCE_PATH.write_text(rendered, encoding='utf-8')
    print(f'Rebuilt {CRANE_BEAM_REFERENCE_PATH.relative_to(REPO_ROOT)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
