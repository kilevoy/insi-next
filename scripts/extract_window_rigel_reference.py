from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
REFERENCE_PATH = REPO_ROOT / 'src/domain/window-rigel/model/window-rigel-reference.generated.ts'

WORKBOOK_ENV_VAR = 'WINDOW_RIGEL_REFERENCE_WORKBOOK'
WORKBOOK_CANDIDATES = [
    Path(os.environ[WORKBOOK_ENV_VAR]) if os.environ.get(WORKBOOK_ENV_VAR) else None,
    Path(r'C:\Users\RAN\Downloads\РџРѕРґР±РѕСЂ РѕРєРѕРЅС‹С… СЂРёРіРµР»РµР№ v2.0.xlsx'),
]

TS_EXPORT_PATTERN = re.compile(r'export const (\w+) = (.*? as const;)', re.S)


def load_snapshot_exports(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}

    text = path.read_text(encoding='utf-8')
    exports: dict[str, Any] = {}

    for export_name, block in TS_EXPORT_PATTERN.findall(text):
        json_text = block[: -len(' as const;')]
        exports[export_name] = json.loads(json_text)

    return exports


def find_workbook_path() -> Path | None:
    for candidate in WORKBOOK_CANDIDATES:
        if candidate and candidate.exists():
            return candidate

    downloads_dir = Path(r'C:\Users\RAN\Downloads')
    if downloads_dir.exists():
        matched = next(downloads_dir.glob('Подбор оконых ригелей v2.0.xlsx'), None)
        if matched and matched.exists():
            return matched

    return None


def extract_candidate_catalog(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for row in range(4, 414):
        if ws[f'R{row}'].value in (None, ''):
            continue

        rows.append(
            {
                'ordinal': int(ws[f'O{row}'].value),
                'excluded': ws[f'N{row}'].value == '-',
                'steelGrade': str(ws[f'P{row}'].value).strip(),
                'strengthResistanceMpa': float(ws[f'Q{row}'].value),
                'profile': str(ws[f'R{row}'].value).strip(),
                'areaCm2': float(ws[f'X{row}'].value),
                'unitMassKgPerM': float(ws[f'Y{row}'].value),
                'momentOfInertiaXCm4': float(ws[f'Z{row}'].value),
                'sectionModulusXCm3': float(ws[f'AA{row}'].value),
                'radiusXcm': float(ws[f'AC{row}'].value),
                'momentOfInertiaYCm4': float(ws[f'AD{row}'].value),
                'sectionModulusYCm3': float(ws[f'AE{row}'].value),
                'radiusYcm': float(ws[f'AF{row}'].value),
            }
        )

    return rows


def extract_window_type_factors(ws) -> list[dict[str, Any]]:
    return [
        {
            'windowType': int(ws[f'I{row}'].value),
            'momentFactor': float(ws[f'J{row}'].value),
            'lengthFactor': float(ws[f'K{row}'].value),
            'deflectionFactor': float(ws[f'L{row}'].value),
        }
        for row in range(34, 39)
    ]


def extract_window_construction_loads(ws) -> list[dict[str, Any]]:
    return [
        {
            'name': str(ws[f'P{row}'].value).strip(),
            'loadKpa': float(ws[f'Q{row}'].value) / 100,
        }
        for row in range(13, 16)
    ]


def extract_pressure_coefficients(ws) -> dict[str, float]:
    return {
        'negativeEdge': float(ws['C25'].value),
        'negativeMiddle': float(ws['C26'].value),
        'positive': float(ws['C35'].value),
        'gustFactor': float(ws['C13'].value),
    }


def extract_terrain_height_table(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for row in range(52, 64):
        rows.append(
            {
                'heightM': int(ws[f'J{row}'].value),
                'kByTerrain': {
                    '\u0410': float(ws[f'K{row}'].value),
                    '\u0412': float(ws[f'L{row}'].value),
                    '\u0421': float(ws[f'M{row}'].value),
                },
                'zetaByTerrain': {
                    '\u0410': float(ws[f'S{row}'].value),
                    '\u0412': float(ws[f'T{row}'].value),
                    '\u0421': float(ws[f'U{row}'].value),
                },
            }
        )

    return rows


def extract_nu_axes_and_grid(ws) -> tuple[list[float], list[float], list[list[float]]]:
    x_values = [float(ws[f'Z{row}'].value) for row in range(68, 75)]
    y_values = [float(ws.cell(row=67, column=column).value) for column in range(27, 34)]
    grid = [
        [float(ws.cell(row=row, column=column).value) for column in range(27, 34)]
        for row in range(68, 75)
    ]

    return x_values, y_values, grid


def extract_exports(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    ws_input = workbook.worksheets[0]
    ws_calc = workbook.worksheets[1]
    ws_wind = workbook.worksheets[3]
    nu_x_values, nu_y_values, nu_grid = extract_nu_axes_and_grid(ws_wind)

    return {
        'windowRigelCandidateCatalog': extract_candidate_catalog(ws_calc),
        'windowRigelWindowTypeFactors': extract_window_type_factors(ws_input),
        'windowRigelWindowConstructionLoads': extract_window_construction_loads(ws_input),
        'windowRigelPressureCoefficients': extract_pressure_coefficients(ws_wind),
        'windowRigelTerrainHeightTable': extract_terrain_height_table(ws_wind),
        'windowRigelNuXValues': nu_x_values,
        'windowRigelNuYValues': nu_y_values,
        'windowRigelNuGrid': nu_grid,
    }


def render_module(exports_map: dict[str, Any], workbook_path: Path | None) -> str:
    if workbook_path is None:
        header = [
            '// Rebuilt from the checked-in window rigel reference snapshot.',
            '// Workbook-backed extraction was skipped because no source workbook was found.',
            '',
        ]
    else:
        header = [
            f'// Rebuilt fully from workbook-backed window rigel references ({workbook_path.name}).',
            '',
        ]

    lines = header
    for export_name, value in exports_map.items():
        serialized = json.dumps(value, ensure_ascii=False, indent=2)
        lines.append(f'export const {export_name} = {serialized} as const;')
        lines.append('')

    return '\n'.join(lines).rstrip() + '\n'


def main() -> int:
    snapshot_exports = load_snapshot_exports(REFERENCE_PATH)
    workbook_path = find_workbook_path()

    if workbook_path is None and not snapshot_exports:
        raise FileNotFoundError('Window rigel workbook not found and no snapshot exists yet.')

    exports_map = extract_exports(workbook_path) if workbook_path is not None else snapshot_exports
    REFERENCE_PATH.write_text(render_module(exports_map, workbook_path), encoding='utf-8')
    print(f'Window rigel references rebuilt from: {workbook_path or REFERENCE_PATH}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
